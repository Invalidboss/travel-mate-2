from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ExcelExportService:
    """Export travel expense data into a pre-formatted workbook template."""

    template_path: Path = Path("templates/Spesenabrechnungsblatt.xlsx")
    mapping_path: Path = Path("backend/config/excel_mapping.yaml")

    def __post_init__(self) -> None:
        self.mapping = self._load_mapping(self.mapping_path)

    @staticmethod
    def _load_mapping(mapping_path: Path) -> dict[str, Any]:
        with mapping_path.open("r", encoding="utf-8") as mapping_file:
            loaded = yaml.safe_load(mapping_file)

        if not isinstance(loaded, dict):
            msg = f"Mapping file must contain a dictionary at root: {mapping_path}"
            raise ValueError(msg)

        return loaded

    def generate_export(self, payload: dict[str, Any], output_path: Path | str) -> Path:
        """Fill template cells based on YAML mapping and save result to output_path."""
        workbook = load_workbook(self.template_path)
        worksheet = workbook[self.mapping["workbook"]["sheet_name"]]

        self._map_meta(worksheet, payload.get("meta", {}))
        self._map_routes(worksheet, payload.get("routes", {}))
        self._map_expenses(worksheet, payload.get("expenses", []))
        self._map_meal_allowance(worksheet, payload.get("meal_allowance", {}))
        self._map_final_totals(worksheet, payload.get("final_totals", {}))

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        return output_path

    def _map_meta(self, sheet: Worksheet, values: dict[str, Any]) -> None:
        for field, cell in self.mapping["meta"].items():
            sheet[cell] = values.get(field)

    def _map_routes(self, sheet: Worksheet, values: dict[str, Any]) -> None:
        routes_map = self.mapping["routes"]
        for route_name, route_cells in routes_map.items():
            route_values = values.get(route_name, {})
            for key, cell in route_cells.items():
                sheet[cell] = route_values.get(key)

    def _map_expenses(self, sheet: Worksheet, values: list[dict[str, Any]]) -> None:
        section = self.mapping["expenses"]
        start_row = int(section["start_row"])
        columns = section["columns"]

        for offset, expense in enumerate(values):
            row = start_row + offset
            for key, column in columns.items():
                sheet[f"{column}{row}"] = expense.get(key)

    def _map_meal_allowance(self, sheet: Worksheet, values: dict[str, Any]) -> None:
        for field, cell in self.mapping["meal_allowance"].items():
            sheet[cell] = values.get(field)

    def _map_final_totals(self, sheet: Worksheet, values: dict[str, Any]) -> None:
        for field, cell in self.mapping["final_totals"].items():
            sheet[cell] = values.get(field)

    def get_mandatory_cells(self) -> list[str]:
        verification = self.mapping.get("verification", {})
        mandatory_cells = verification.get("mandatory_cells", [])
        if not isinstance(mandatory_cells, list):
            msg = "verification.mandatory_cells must be a list of cell references"
            raise ValueError(msg)
        return mandatory_cells


def read_cells(path: Path | str, cells: list[str], sheet_name: str) -> dict[str, Any]:
    """Utility for validation/testing: read exact cell values from an exported workbook."""
    workbook: Workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name]
    return {cell: sheet[cell].value for cell in cells}
