from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

from backend.services.excel_export import ExcelExportService, read_cells


def ensure_template(template_path: Path, sheet_name: str) -> None:
    """Create a minimal demo template only when the expected template is missing."""
    if template_path.exists():
        return

    template_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.merge_cells("A1:D1")
    ws["A1"] = "Spesenabrechnungsblatt"
    ws["A1"].font = Font(bold=True, size=14)

    thin = Side(border_style="thin", color="000000")
    for row in range(2, 35):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws["C28"] = "Subtotal"
    ws["D28"] = "=SUM(D12:D21)"
    ws["C29"] = "Grand total"
    ws["D29"] = "=D28+D25"

    wb.save(template_path)


def main() -> int:
    service = ExcelExportService()
    sheet_name = service.mapping["workbook"]["sheet_name"]

    ensure_template(service.template_path, sheet_name)

    payload = {
        "meta": {
            "employee_name": "Max Mustermann",
            "project_name": "Project Alpen",
            "trip_period": "2026-02-01 to 2026-02-05",
        },
        "routes": {
            "outbound": {"from": "Berlin", "to": "Munich", "date": "2026-02-01"},
            "return": {"from": "Munich", "to": "Berlin", "date": "2026-02-05"},
            "connecting": {"from": "Munich", "to": "Innsbruck", "date": "2026-02-03"},
        },
        "expenses": [
            {
                "category": "Hotel",
                "account": "6600",
                "description": "4 nights",
                "amount": 560.0,
            },
            {
                "category": "Transport",
                "account": "6670",
                "description": "Train tickets",
                "amount": 120.0,
            },
        ],
        "meal_allowance": {
            "breakfast_days": 2,
            "lunch_days": 3,
            "dinner_days": 3,
            "total": 96.0,
        },
        "final_totals": {
            "subtotal_expenses": 680.0,
            "grand_total": 776.0,
        },
    }

    output_path = Path("artifacts/sample_expense_export.xlsx")
    service.generate_export(payload, output_path)

    mandatory_cells = service.get_mandatory_cells()
    values = read_cells(output_path, mandatory_cells, sheet_name)
    missing = [cell for cell, value in values.items() if value in (None, "")]

    if missing:
        print("Verification failed. Missing mandatory values in:", ", ".join(missing))
        return 1

    print(f"Verification passed. Export generated at {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
